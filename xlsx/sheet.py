from logging import exception
import openpyxl

path=str('quiz.xlsx')

try:
    file_obj=openpyxl.load_workbook(path.strip())
    sheet_obj=file_obj.active

    max_column= sheet_obj.max_column
    max_row= sheet_obj.max_row

    # To get all the details of the sheet at a time
    for row in range(2,max_row+1):
       for column in range(1,max_column+1):
           c=sheet_obj.cell(row=row,column=column)
           print(c.value,end="   ")
       print(" ")

    #To modify any cell data
    c1=sheet_obj.cell(row=2,column=2)
    c1.value="Hello"
    print(c1.value)

    #To duplicate this xlsx file
    file_obj.save("Modified.xlsx")
    
except exception as e:
    print(e)
    print("Error")
    